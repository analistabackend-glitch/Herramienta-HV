"""
generador_excel_unificado.py
============================
Genera un único Excel consolidado con tres hojas:

  Hoja 1 – "Resumen HV descargadas"   : datos del primer filtro (TODOS los candidatos)
  Hoja 2 – "Resumen resultados finales": datos del tercer filtro (candidatos evaluados por IA)
  Hoja 3 – "Parámetros"               : parámetros del filtro 1 y del filtro 3

El archivo se guarda dentro de la carpeta de opcionados del tercer filtro:
  Resultados Tercer Filtro/<vacante>_<mes>/resumen_completo_<vacante>.xlsx
"""

import os
import json
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────
# COLORES Y ESTILOS
# ─────────────────────────────────────────────────────────

COLOR_NARANJA     = "FF9900"   # encabezados principales
COLOR_NARANJA_CLR = "FFF3E0"   # fila par / relleno suave
COLOR_VERDE       = "E8F5E9"   # aceptado / opcionado
COLOR_ROJO        = "FFEBEE"   # rechazado / descartado
COLOR_AMARILLO    = "FFFDE7"   # probable
COLOR_GRIS        = "F5F5F5"   # sin pdf / fila par
COLOR_BLANCO      = "FFFFFF"
COLOR_NARANJA_ADV = "FFF3CD"   # advertencia: documento no HV

ESTADOS_COLOR = {
    # primer filtro
    "SUBIDO"             : COLOR_VERDE,
    "RECHAZADO"          : COLOR_ROJO,
    "SIN PDF"            : COLOR_GRIS,
    "PDF NO DESCARGABLE" : COLOR_GRIS,
    # tercer filtro
    "Opcionados"               : COLOR_VERDE,
    "Probablemente Opcionados" : COLOR_AMARILLO,
    "Descartados"              : COLOR_ROJO,
    # advertencia
    "⚠️ NO ES HV"              : COLOR_NARANJA_ADV,
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border_thin() -> Border:
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _estilo_encabezado(ws, fila: int):
    """Aplica estilo de encabezado naranja a toda la fila indicada."""
    for cell in ws[fila]:
        cell.fill      = _fill(COLOR_NARANJA)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border_thin()


def _autofit(ws):
    """Ajusta el ancho de columnas al contenido (aproximado)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                largo = len(str(cell.value)) if cell.value else 0
                if largo > max_len:
                    max_len = largo
            except Exception:
                pass
        # mínimo 12, máximo 60
        ws.column_dimensions[col_letter].width = max(12, min(60, max_len + 4))


def _color_fila_estado(estado: str) -> str:
    return ESTADOS_COLOR.get(estado, COLOR_BLANCO)


# ─────────────────────────────────────────────────────────
# HOJA 1 — RESUMEN HV DESCARGADAS  (primer filtro)
# ─────────────────────────────────────────────────────────

def _construir_hoja1(writer, resumen_primer_filtro: list):
    """
    Escribe la hoja 'Resumen HV descargadas'.
    Cada elemento de resumen_primer_filtro es un dict con las claves
    que genera correr_proceso en primer_filtro.py.
    """
    filas = []
    for d in resumen_primer_filtro:
        estado = d.get("estado", "")
        # Motivo: si fue aceptado usamos motivo_seleccion, si no motivo_rechazo
        if estado == "SUBIDO":
            motivo = d.get("motivo_seleccion", "")
        else:
            motivo = d.get("motivo_rechazo", "")

        filas.append({
            "Estado"           : estado,
            "Motivo"           : motivo,
            "Nombre"           : d.get("nombre", ""),
            "Edad"             : d.get("edad", ""),
            "Salario aspirado" : d.get("salario", ""),
            "Sábados"          : d.get("sabados", ""),
            "URL perfil"       : d.get("url", ""),
        })

    # Ordenar: aceptados primero
    orden = {"SUBIDO": 0, "SIN PDF": 1, "PDF NO DESCARGABLE": 2, "RECHAZADO": 3}
    filas.sort(key=lambda x: orden.get(x["Estado"], 4))

    df = pd.DataFrame(filas) if filas else pd.DataFrame(
        columns=["Estado", "Motivo", "Nombre", "Edad", "Salario aspirado", "Sábados", "URL perfil"]
    )
    df.to_excel(writer, index=False, sheet_name="Resumen HV descargadas")

    ws = writer.sheets["Resumen HV descargadas"]

    # Encabezado
    _estilo_encabezado(ws, 1)
    ws.row_dimensions[1].height = 30

    # Filas de datos
    for row_idx, fila in enumerate(filas, start=2):
        color = _color_fila_estado(fila["Estado"])
        for cell in ws[row_idx]:
            cell.fill      = _fill(color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border_thin()

    # Congelar primera fila
    ws.freeze_panes = "A2"
    _autofit(ws)


# ─────────────────────────────────────────────────────────
# HOJA 2 — RESUMEN RESULTADOS FINALES  (tercer filtro)
# ─────────────────────────────────────────────────────────

def _construir_hoja2(writer, resumen_tercer_filtro: list):
    """
    Escribe la hoja 'Resumen resultados finales'.
    Garantiza que SIEMPRE se llenen ambos análisis (experiencia y académico),
    aun cuando el candidato fue descartado por inestabilidad laboral.
    """
    columnas = [
        "Candidato",
        "Clasificación",
        "Score Final (%)",
        "Score Experiencia (%)",
        "Score Académico (%)",
        "Razón Experiencia",
        "Razón Académica",
        "Resumen",
        "Archivo",
    ]

    filas = []
    for r in resumen_tercer_filtro:
        # ── Garantizar que ambos análisis estén completos ──────────────────
        razon_exp = r.get("Razón Experiencia") or ""
        razon_aca = r.get("Razón Académica")   or ""

        # Si uno está vacío, usar el resumen como fallback descriptivo
        resumen_general = r.get("Resumen") or ""
        if not razon_exp and resumen_general:
            razon_exp = f"(Ver resumen general) {resumen_general}"
        if not razon_aca and resumen_general:
            razon_aca = f"(Ver resumen general) {resumen_general}"

        # Marcar celdas sin score como N/A cuando score = 0 por error, no por evaluación
        score_exp = r.get("Score Experiencia (%)") or 0
        score_aca = r.get("Score Académico (%)")   or 0

        filas.append({
            "Candidato"             : r.get("Candidato", ""),
            "Clasificación"         : r.get("Clasificación", ""),
            "Score Final (%)"       : r.get("Score Final (%)", 0),
            "Score Experiencia (%)" : score_exp,
            "Score Académico (%)"   : score_aca,
            "Razón Experiencia"     : razon_exp,
            "Razón Académica"       : razon_aca,
            "Resumen"               : resumen_general,
            "Archivo"               : r.get("Archivo", ""),
        })

    df = pd.DataFrame(filas) if filas else pd.DataFrame(columns=columnas)
    df.to_excel(writer, index=False, sheet_name="Resumen resultados finales")

    ws = writer.sheets["Resumen resultados finales"]

    _estilo_encabezado(ws, 1)
    ws.row_dimensions[1].height = 30

    for row_idx, fila in enumerate(filas, start=2):
        # Detectar si es un documento no-HV por la razón de experiencia
        es_no_hv = str(fila.get("Razón Experiencia", "")).startswith("⚠️ DOCUMENTO NO CORRESPONDE")
        if es_no_hv:
            color = COLOR_NARANJA_ADV
        else:
            color = _color_fila_estado(fila["Clasificación"])
        for cell in ws[row_idx]:
            cell.fill      = _fill(color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border_thin()
        ws.row_dimensions[row_idx].height = 50

    ws.freeze_panes = "A2"
    _autofit(ws)


# ─────────────────────────────────────────────────────────
# HOJA 3 — PARÁMETROS (filtro 1 + filtro 3)
# ─────────────────────────────────────────────────────────

def _construir_hoja3(writer, cfg_filtro1: dict, vacante_data: dict,
                     umbral_opcionado: int, umbral_probable: int,
                     conteos: dict):
    """
    Escribe la hoja 'Parámetros' con todos los parámetros de ambos filtros.
    """
    filas = [
        # ── Bloque: Vacante ──────────────────────────────────────────────
        {"Sección": "Vacante", "Parámetro": "Nombre de la vacante",
         "Valor": cfg_filtro1.get("vacante", vacante_data.get("vacante", ""))},
        {"Sección": "Vacante", "Parámetro": "URL vacante",
         "Valor": cfg_filtro1.get("url_vacante", "")},

        # ── Bloque: Filtro 1 ─────────────────────────────────────────────
        {"Sección": "Filtro 1 — Descarga y filtrado básico",
         "Parámetro": "Edad mínima",
         "Valor": cfg_filtro1.get("edad_min", "")},
        {"Sección": "Filtro 1 — Descarga y filtrado básico",
         "Parámetro": "Edad máxima",
         "Valor": cfg_filtro1.get("edad_max", "")},
        {"Sección": "Filtro 1 — Descarga y filtrado básico",
         "Parámetro": "Salario mínimo ($)",
         "Valor": f"${cfg_filtro1['sal_min']:,}" if cfg_filtro1.get("sal_min") else ""},
        {"Sección": "Filtro 1 — Descarga y filtrado básico",
         "Parámetro": "Salario máximo ($)",
         "Valor": f"${cfg_filtro1['sal_max']:,}" if cfg_filtro1.get("sal_max") else ""},
        {"Sección": "Filtro 1 — Descarga y filtrado básico",
         "Parámetro": "Requiere disponibilidad sábados",
         "Valor": "Sí" if cfg_filtro1.get("requiere_sabados") else "No"},

        # ── Bloque: Pesos de evaluación ──────────────────────────────────
        {"Sección": "Pesos de evaluación (compartido filtros 1 y 3)",
         "Parámetro": "Peso experiencia laboral",
         "Valor": vacante_data.get("peso_experiencia_laboral",
                  f"{cfg_filtro1.get('peso_exp', '')} %")},
        {"Sección": "Pesos de evaluación (compartido filtros 1 y 3)",
         "Parámetro": "Peso formación académica",
         "Valor": vacante_data.get("peso_formacion_academica",
                  f"{cfg_filtro1.get('peso_aca', '')} %")},

        # ── Bloque: Filtro 3 ─────────────────────────────────────────────
        {"Sección": "Filtro 3 — Scoring IA",
         "Parámetro": "Umbral Opcionado (score ≥)",
         "Valor": f"{umbral_opcionado} %"},
        {"Sección": "Filtro 3 — Scoring IA",
         "Parámetro": "Umbral Probablemente Opcionado (score ≥)",
         "Valor": f"{umbral_probable} %"},
        {"Sección": "Filtro 3 — Scoring IA",
         "Parámetro": "Umbral Descartado (score <)",
         "Valor": f"{umbral_probable} %"},

        # ── Bloque: Resultados ───────────────────────────────────────────
        {"Sección": "Resultados del proceso",
         "Parámetro": "Total candidatos revisados (filtro 1)",
         "Valor": conteos.get("total_f1", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Candidatos aceptados (filtro 1)",
         "Valor": conteos.get("aceptados_f1", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Candidatos rechazados (filtro 1)",
         "Valor": conteos.get("rechazados_f1", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Candidatos sin PDF",
         "Valor": conteos.get("sin_pdf_f1", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Total evaluados (filtro 3)",
         "Valor": conteos.get("total_f3", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Opcionados",
         "Valor": conteos.get("opcionados", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Probablemente Opcionados",
         "Valor": conteos.get("probables", "")},
        {"Sección": "Resultados del proceso",
         "Parámetro": "Descartados (filtro 3)",
         "Valor": conteos.get("descartados", "")},

        {"Sección": "Resultados del proceso",
         "Parámetro": "Documentos no identificados como HV (verificar)",
         "Valor": conteos.get("no_hv", 0)},

        # ── Bloque: Metadatos ────────────────────────────────────────────
        {"Sección": "Metadatos",
         "Parámetro": "Fecha de ejecución",
         "Valor": datetime.now().strftime("%Y-%m-%d %H:%M")},
    ]

    df = pd.DataFrame(filas)
    df.to_excel(writer, index=False, sheet_name="Parámetros")

    ws = writer.sheets["Parámetros"]
    _estilo_encabezado(ws, 1)
    ws.row_dimensions[1].height = 28

    # Colorear secciones alternadamente y en negrita la sección
    seccion_actual = None
    color_alt      = [COLOR_NARANJA_CLR, COLOR_BLANCO]
    idx_color      = 0

    for row_idx in range(2, ws.max_row + 1):
        seccion_cell = ws.cell(row=row_idx, column=1)
        if seccion_cell.value and seccion_cell.value != seccion_actual:
            seccion_actual = seccion_cell.value
            idx_color = (idx_color + 1) % 2

        color = color_alt[idx_color]
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = _fill(color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border_thin()

        # Negrita en la columna "Sección" solo cuando cambia
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=9)
        ws.cell(row=row_idx, column=2).font = Font(size=9)
        ws.cell(row=row_idx, column=3).font = Font(size=9)

    ws.freeze_panes = "A2"
    _autofit(ws)


# ─────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL EXPORTADA
# ─────────────────────────────────────────────────────────

def generar_excel_unificado(
    resumen_primer_filtro: list,
    resumen_tercer_filtro: list,
    cfg_filtro1: dict,
    vacante_data: dict,
    carpeta_salida: Path,
    umbral_opcionado: int = 70,
    umbral_probable: int  = 45,
    archivos_no_hv: list  = None,
) -> Path:
    """
    Genera el Excel consolidado con las 3 hojas y lo guarda en carpeta_salida.

    Parámetros
    ----------
    resumen_primer_filtro : lista de dicts producida en correr_proceso() de primer_filtro.py
    resumen_tercer_filtro : lista de dicts producida en main() de tercer_filtro.py
    cfg_filtro1           : dict con los parámetros del primer filtro (edad, salario, pesos…)
    vacante_data          : dict leído del descripcion_*.json (peso exp/aca, nombre vacante…)
    carpeta_salida        : Path donde se guardará el Excel (carpeta base del tercer filtro)
    umbral_opcionado      : puntaje mínimo para ser Opcionado
    umbral_probable       : puntaje mínimo para ser Probablemente Opcionado

    Retorna
    -------
    Path al archivo Excel generado.
    """
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    nombre_vacante = cfg_filtro1.get("vacante") or vacante_data.get("vacante", "vacante")
    nombre_limpio  = re.sub(r"[^\w\s-]", "", nombre_vacante).strip().replace(" ", "_")[:60]
    excel_path     = carpeta_salida / f"resumen_completo_{nombre_limpio}.xlsx"

    archivos_no_hv = archivos_no_hv or []

    # ── Conteos para hoja de parámetros ────────────────────────────────
    conteos = {
        "total_f1"      : len(resumen_primer_filtro),
        "aceptados_f1"  : sum(1 for d in resumen_primer_filtro if d.get("estado") == "SUBIDO"),
        "rechazados_f1" : sum(1 for d in resumen_primer_filtro if d.get("estado") == "RECHAZADO"),
        "sin_pdf_f1"    : sum(1 for d in resumen_primer_filtro
                              if d.get("estado") in ("SIN PDF", "PDF NO DESCARGABLE")),
        "total_f3"      : len(resumen_tercer_filtro),
        "opcionados"    : sum(1 for r in resumen_tercer_filtro
                              if r.get("Clasificación") == "Opcionados"),
        "probables"     : sum(1 for r in resumen_tercer_filtro
                              if r.get("Clasificación") == "Probablemente Opcionados"),
        "descartados"   : sum(1 for r in resumen_tercer_filtro
                              if r.get("Clasificación") == "Descartados"),
        "no_hv"         : len(archivos_no_hv),
    }

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        _construir_hoja1(writer, resumen_primer_filtro)
        _construir_hoja2(writer, resumen_tercer_filtro)
        _construir_hoja3(writer, cfg_filtro1, vacante_data,
                         umbral_opcionado, umbral_probable, conteos)

    print(f"✅ Excel unificado generado: {excel_path}")
    return excel_path
